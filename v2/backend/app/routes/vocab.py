import re
from io import BytesIO

from flask import Blueprint, request
from flask_jwt_extended import get_jwt_identity, jwt_required
from openpyxl import load_workbook

from .. import db
from ..models import ImportBatch, UserCard, User, Vocabulary, VocabularyWord, Word

bp = Blueprint("vocab", __name__)


def normalize_word(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"\s+", " ", value)
    return value


def row_to_word(row):
    values = [str(v).strip() if v is not None else "" for v in row]
    if not values or not values[0]:
        return None
    # V2 importer accepts: word, meaning, example, translation, source.
    return {
        "word": values[0],
        "meaning": values[1] if len(values) > 1 else "",
        "example": values[2] if len(values) > 2 else "",
        "translation": values[3] if len(values) > 3 else "",
        "source": values[4] if len(values) > 4 else "",
    }


@bp.post("/import")
@jwt_required()
def import_vocab():
    user_id = get_jwt_identity()
    file = request.files.get("file")
    if not file:
        return {"message": "缺少文件"}, 400
    name = file.filename or "import"
    if name.lower().endswith(".csv"):
        import csv
        import io
        rows = list(csv.reader(io.TextIOWrapper(file.stream, encoding="utf-8-sig")))
    elif name.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    else:
        return {"message": "目前支持 CSV、XLSX"}, 415

    batch = ImportBatch(user_id=user_id, filename=name, total_rows=max(0, len(rows) - 1))
    db.session.add(batch)
    vocabulary = Vocabulary(owner_user_id=user_id, name=name.rsplit(".", 1)[0], kind="user")
    db.session.add(vocabulary)
    db.session.flush()

    seen = set()
    for index, raw in enumerate(rows[1:], start=2):
        item = row_to_word(raw)
        if not item:
            batch.invalid_rows += 1
            continue
        normalized = normalize_word(item["word"])
        if normalized in seen:
            batch.duplicates_in_file += 1
            continue
        seen.add(normalized)
        batch.valid_rows += 1
        word = Word.query.filter_by(normalized=normalized).first()
        if word:
            batch.linked_existing += 1
        else:
            word = Word(normalized=normalized, word=item["word"], meaning=item["meaning"])
            db.session.add(word)
            db.session.flush()
            batch.inserted_words += 1
        link = VocabularyWord.query.filter_by(vocabulary_id=vocabulary.id, word_id=word.id).first()
        if not link:
            db.session.add(VocabularyWord(vocabulary_id=vocabulary.id, word_id=word.id, note=item["example"]))

    db.session.commit()
    return {"batch_id": batch.id, "vocabulary_id": vocabulary.id, "total_rows": batch.total_rows, "valid_rows": batch.valid_rows, "inserted_words": batch.inserted_words, "linked_existing": batch.linked_existing, "duplicates_in_file": batch.duplicates_in_file, "invalid_rows": batch.invalid_rows}, 201


@bp.get("")
@jwt_required()
def list_vocabularies():
    user_id = get_jwt_identity()
    rows = Vocabulary.query.filter((Vocabulary.owner_user_id == user_id) | (Vocabulary.kind == "system")).all()
    return [{"id": v.id, "name": v.name, "kind": v.kind} for v in rows]
