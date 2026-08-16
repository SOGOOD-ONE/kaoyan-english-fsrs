import re
import unicodedata
from openpyxl import load_workbook
from .. import db
from ..models import Word


def normalize_word(value: str) -> str:
    value = unicodedata.normalize('NFKC', str(value or '')).strip().casefold()
    value = re.sub(r'\s+', ' ', value)
    return value


def import_xlsx(file_obj):
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'total': 0, 'inserted': 0, 'merged': 0, 'invalid': 0, 'duplicate': 0}
    headers = [str(x or '').strip().casefold() for x in rows[0]]
    def idx(*names):
        for name in names:
            if name in headers:
                return headers.index(name)
        return None
    wi, mi = idx('word', '单词', '词汇'), idx('meaning', '释义', '中文')
    ei, ti = idx('example', '例句', 'sentence'), idx('translation', '翻译', '译文')
    if wi is None:
        raise ValueError('Excel 必须包含 word / 单词 / 词汇 列')
    result = {'total': max(0, len(rows)-1), 'inserted': 0, 'merged': 0, 'invalid': 0, 'duplicate': 0}
    seen = set()
    for row in rows[1:]:
        raw = row[wi] if wi < len(row) else None
        key = normalize_word(raw)
        if not key:
            result['invalid'] += 1
            continue
        if key in seen:
            result['duplicate'] += 1
            continue
        seen.add(key)
        existing = db.session.query(Word).filter_by(normalized_word=key).first()
        if existing:
            result['merged'] += 1
            continue
        word = Word(word=str(raw).strip(), normalized_word=key,
                    meaning=str(row[mi]).strip() if mi is not None and row[mi] is not None else '',
                    example=str(row[ei]).strip() if ei is not None and row[ei] is not None else None,
                    translation=str(row[ti]).strip() if ti is not None and row[ti] is not None else None,
                    source='import')
        db.session.add(word)
        result['inserted'] += 1
    db.session.commit()
    return result
