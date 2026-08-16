"""Import an Excel vocabulary workbook into the system vocabulary tables.

Expected columns (order-independent): 单词, 词性, 释义, 分类.
Optional columns are ignored safely. The importer is idempotent by normalized word.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app import create_app, db  # noqa: E402
from app.models import Vocabulary, VocabularyWord, Word  # noqa: E402

REQUIRED = {"单词", "词性", "释义", "分类"}
HEADER_ALIASES = {
    "word": {"单词", "词汇", "word", "english", "英文"},
    "type": {"词性", "词性/短语", "part of speech", "pos"},
    "meaning": {"释义", "中文释义", "意思", "meaning"},
    "category": {"分类", "类别", "category", "类型"},
}


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def clean(value: object) -> str:
    return str(value or "").strip()


def find_header(rows):
    for index, row in enumerate(rows, start=1):
        names = {clean(v) for v in row}
        if REQUIRED.issubset(names):
            return index, row
    raise ValueError(f"找不到包含 {', '.join(sorted(REQUIRED))} 的表头")


def column_map(header):
    result = {}
    for index, value in enumerate(header):
        name = clean(value).lower()
        for key, aliases in HEADER_ALIASES.items():
            if name in {a.lower() for a in aliases}:
                result[key] = index
    missing = {"word", "type", "meaning", "category"} - set(result)
    if missing:
        raise ValueError(f"缺少列: {', '.join(sorted(missing))}")
    return result


def import_workbook(path: Path, vocabulary_name: str, priority: int, source: str) -> tuple[int, int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    app = create_app()
    inserted = updated = skipped = 0
    with app.app_context():
        vocabulary = Vocabulary.query.filter_by(name=vocabulary_name, kind="system").first()
        if not vocabulary:
            vocabulary = Vocabulary(name=vocabulary_name, kind="system", priority=priority, description=f"Excel导入：{path.name}")
            db.session.add(vocabulary)
            db.session.flush()
        else:
            vocabulary.priority = priority

        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            buffered = []
            try:
                header_index, header = find_header([next(rows) for _ in range(0)])
            except ValueError:
                # The workbook may have title rows; read a bounded prefix instead.
                sheet_rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True))
                try:
                    header_index, header = find_header(sheet_rows)
                except ValueError:
                    continue
                rows = sheet.iter_rows(min_row=header_index + 1, values_only=True)
            columns = column_map(header)
            for row in rows:
                word_value = clean(row[columns["word"]] if columns["word"] < len(row) else "")
                if not word_value:
                    continue
                normalized = normalize(word_value)
                if not normalized:
                    skipped += 1
                    continue
                values = {
                    "word_type": clean(row[columns["type"]] if columns["type"] < len(row) else ""),
                    "meaning": clean(row[columns["meaning"]] if columns["meaning"] < len(row) else ""),
                    "category": clean(row[columns["category"]] if columns["category"] < len(row) else ""),
                }
                existing = Word.query.filter_by(normalized_word=normalized).first()
                if existing:
                    # Keep the canonical spelling but refresh the metadata from the source workbook.
                    existing.word_type = values["word_type"]
                    existing.meaning = values["meaning"]
                    existing.category = values["category"]
                    existing.source = source
                    existing.source_detail = path.name
                    word = existing
                    updated += 1
                else:
                    word = Word(word=word_value, normalized_word=normalized, source=source, source_detail=path.name, **values)
                    db.session.add(word)
                    db.session.flush()
                    inserted += 1
                if not VocabularyWord.query.filter_by(vocabulary_id=vocabulary.id, word_id=word.id).first():
                    db.session.add(VocabularyWord(vocabulary_id=vocabulary.id, word_id=word.id, priority=priority))
        db.session.commit()
    workbook.close()
    return inserted, updated, skipped


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("file", type=Path)
    parser.add_argument("--name", default="考研英语核心词")
    parser.add_argument("--priority", type=int, default=100)
    parser.add_argument("--source", default="excel")
    args = parser.parse_args()
    if not args.file.exists():
        raise SystemExit(f"文件不存在: {args.file}")
    result = import_workbook(args.file, args.name, args.priority, args.source)
    print(f"导入完成：新增 {result[0]}，更新 {result[1]}，跳过 {result[2]}")


if __name__ == "__main__":
    main()
